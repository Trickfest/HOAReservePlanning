from __future__ import annotations

import math
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .model import Component, Inputs, compute_forecast, expenses_by_year
from .schedule import ScheduleItem

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9D9D9")
INPUT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
BOLD_FONT = Font(bold=True)

INPUT_ROWS = {
    "starting_year": 2,
    "beginning_reserve_balance": 3,
    "inflation_rate": 4,
    "investment_return_rate": 5,
    "spend_inflation_timing": 6,
    "audit_tolerance_amount": 7,
    "audit_tolerance_ratio": 8,
}
FEATURE_ROWS = {
    "forecast_years": 11,
    "enable_checks": 12,
    "enable_dashboard": 13,
    "enable_schedule_expansion": 14,
    "enable_audit": 15,
    "max_components_rows": 16,
    "max_schedule_rows": 17,
}

FORECAST_HEADERS = [
    "year",
    "begin_balance",
    "contributions",
    "cumulative_contributions",
    "interest",
    "cumulative_interest",
    "expenses",
    "end_balance",
    "percent_funded",
    "coverage_5yr",
]
AUDIT_RATIO_HEADERS = {"percent_funded", "coverage_5yr"}


def _format_inflation_offset(offset: float) -> str:
    if offset == 0:
        return ""
    if offset.is_integer():
        return f"+{int(offset)}"
    return f"+{offset}"


def build_workbook(
    inputs: Inputs,
    components: List[Component],
    schedule_items: List[ScheduleItem],
    contributions: Dict[int, float],
    scenario: str,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    _write_readme_sheet(ws, scenario)

    inputs_ws = wb.create_sheet("Inputs")
    _write_inputs_sheet(inputs_ws, inputs)

    components_ws = wb.create_sheet("Components")
    _write_components_sheet(components_ws, inputs, components)

    schedule_ws = wb.create_sheet("Schedule")
    _write_schedule_sheet(schedule_ws, inputs, schedule_items)

    funding_ws = wb.create_sheet("Funding")
    _write_funding_sheet(funding_ws, inputs)

    forecast_ws = wb.create_sheet("Forecast")
    _write_forecast_sheet(forecast_ws, inputs, components, schedule_items, contributions)

    if inputs.features.get("enable_checks", True):
        checks_ws = wb.create_sheet("Checks")
        _write_checks_sheet(checks_ws, inputs)
    else:
        checks_ws = wb.create_sheet("Checks")
        checks_ws["A1"] = "Checks disabled by FEATURES.enable_checks"

    if inputs.features.get("enable_dashboard", True):
        dashboard_ws = wb.create_sheet("Dashboard")
        _write_dashboard_sheet(dashboard_ws, inputs)
    else:
        dashboard_ws = wb.create_sheet("Dashboard")
        dashboard_ws["A1"] = "Dashboard disabled by FEATURES.enable_dashboard"

    return wb


def _style_header(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = BOLD_FONT
        cell.alignment = Alignment(horizontal="left")


def _write_readme_sheet(ws, scenario: str) -> None:
    summary_lines = [
        "HOA Reserve Planning Workbook",
        f"Scenario: {scenario}",
        "",
        "Generated from:",
        "- data/inputs.yaml",
        "- data/components.csv",
        "- data/contributions/*.csv",
        "",
        "Edit source files and rebuild for versioned changes.",
        "",
        "Sheets: Inputs, Components, Schedule, Forecast, Checks, Dashboard",
    ]
    ws["A1"] = "\n".join(summary_lines)
    ws["A1"].alignment = Alignment(vertical="top", wrap_text=True)
    ws["A1"].fill = HEADER_FILL

    for row in range(2, 20):
        ws.row_dimensions[row].hidden = True
    ws.column_dimensions["A"].width = 90
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 18 * len(summary_lines)


def _write_inputs_sheet(ws, inputs: Inputs) -> None:
    ws.append(["Input", "Value", "Notes"])
    _style_header(ws, 3)

    rows = [
        ("starting_year", inputs.starting_year, "First year of the forecast"),
        ("beginning_reserve_balance", inputs.beginning_reserve_balance, "Opening reserve balance"),
        ("inflation_rate", inputs.inflation_rate, "Annual inflation rate"),
        ("investment_return_rate", inputs.investment_return_rate, "Annual investment return"),
        ("spend_inflation_timing", inputs.spend_inflation_timing, "Timing for spend inflation (start_of_year, mid_year, end_of_year)"),
        ("audit_tolerance_amount", inputs.audit_tolerance_amount, "Audit tolerance for dollar values"),
        ("audit_tolerance_ratio", inputs.audit_tolerance_ratio, "Audit tolerance for ratios (percent funded, coverage)"),
    ]

    for label, value, note in rows:
        ws.append([label, value, note])

    ws.append(["", "", ""])
    ws.append(["FEATURES", "", "Toggle optional behavior"])
    ws.append(["forecast_years", inputs.forecast_years, "Forecast horizon in years"])
    ws.append([
        "enable_checks",
        bool(inputs.features.get("enable_checks", True)),
        "Include the Checks sheet",
    ])
    ws.append([
        "enable_dashboard",
        bool(inputs.features.get("enable_dashboard", True)),
        "Include the Dashboard sheet",
    ])
    ws.append([
        "enable_schedule_expansion",
        bool(inputs.features.get("enable_schedule_expansion", True)),
        "Generate recurring schedule rows",
    ])
    ws.append([
        "enable_audit",
        bool(inputs.features.get("enable_audit", False)),
        "Add audit columns and summary checks",
    ])
    ws.append([
        "max_components_rows",
        int(inputs.features.get("max_components_rows", 500)),
        "Max rows reserved for Components",
    ])
    ws.append([
        "max_schedule_rows",
        int(inputs.features.get("max_schedule_rows", 10000)),
        "Max rows reserved for Schedule",
    ])

    for row in sorted(INPUT_ROWS.values()):
        ws.cell(row=row, column=2).fill = INPUT_FILL

    for row in sorted(FEATURE_ROWS.values()):
        ws.cell(row=row, column=2).fill = INPUT_FILL

    ws.cell(row=INPUT_ROWS["starting_year"], column=2).number_format = "0"
    ws.cell(row=INPUT_ROWS["beginning_reserve_balance"], column=2).number_format = "#,##0"
    ws.cell(row=INPUT_ROWS["inflation_rate"], column=2).number_format = "0.00%"
    ws.cell(row=INPUT_ROWS["investment_return_rate"], column=2).number_format = "0.00%"
    ws.cell(row=INPUT_ROWS["spend_inflation_timing"], column=2).number_format = "@"
    ws.cell(row=INPUT_ROWS["audit_tolerance_amount"], column=2).number_format = "#,##0.00"
    ws.cell(row=INPUT_ROWS["audit_tolerance_ratio"], column=2).number_format = "0.00%"

    ws.cell(row=FEATURE_ROWS["forecast_years"], column=2).number_format = "0"
    ws.cell(row=FEATURE_ROWS["max_components_rows"], column=2).number_format = "0"
    ws.cell(row=FEATURE_ROWS["max_schedule_rows"], column=2).number_format = "0"

    summary_lines = [
        "Inputs",
        f"starting_year: {inputs.starting_year}",
        f"beginning_reserve_balance: {inputs.beginning_reserve_balance:,.0f}",
        f"inflation_rate: {inputs.inflation_rate:.2%}",
        f"investment_return_rate: {inputs.investment_return_rate:.2%}",
        f"spend_inflation_timing: {inputs.spend_inflation_timing}",
        f"audit_tolerance_amount: {inputs.audit_tolerance_amount:,.2f}",
        f"audit_tolerance_ratio: {inputs.audit_tolerance_ratio:.4%}",
        "",
        "Features",
        f"forecast_years: {inputs.forecast_years}",
        f"enable_checks: {str(bool(inputs.features.get('enable_checks', True))).lower()}",
        f"enable_dashboard: {str(bool(inputs.features.get('enable_dashboard', True))).lower()}",
        f"enable_schedule_expansion: {str(bool(inputs.features.get('enable_schedule_expansion', True))).lower()}",
        f"enable_audit: {str(bool(inputs.features.get('enable_audit', False))).lower()}",
        f"max_components_rows: {int(inputs.features.get('max_components_rows', 500))}",
        f"max_schedule_rows: {int(inputs.features.get('max_schedule_rows', 10000))}",
    ]
    ws["A1"] = "\n".join(summary_lines)
    ws["A1"].alignment = Alignment(vertical="top", wrap_text=True)

    # Hide the input table while keeping values for formulas.
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].hidden = True
    ws.column_dimensions["B"].hidden = True
    ws.column_dimensions["C"].hidden = True
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 18 * len(summary_lines)
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 90


def _write_components_sheet(ws, inputs: Inputs, components: List[Component]) -> None:
    headers = [
        "id",
        "name",
        "category",
        "base_cost",
        "spend_year",
        "recurring",
        "interval_years",
        "include",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    max_rows = int(inputs.features.get("max_components_rows", 500))
    for row in range(2, 2 + max_rows):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = INPUT_FILL

    for component in components:
        row = component.row_index
        ws.cell(row=row, column=1, value=component.id)
        ws.cell(row=row, column=2, value=component.name)
        ws.cell(row=row, column=3, value=component.category)
        ws.cell(row=row, column=4, value=component.base_cost)
        ws.cell(row=row, column=5, value=component.spend_year)
        ws.cell(row=row, column=6, value="Y" if component.recurring else "N")
        ws.cell(row=row, column=7, value=component.interval_years or "")
        ws.cell(row=row, column=8, value="Y" if component.include else "N")

    for row in range(2, 2 + max_rows):
        ws.cell(row=row, column=4).number_format = "#,##0"
        ws.cell(row=row, column=5).number_format = "0"
        ws.cell(row=row, column=7).number_format = "0"

    ws.freeze_panes = "A2"
    widths = [16, 28, 18, 14, 12, 10, 14, 10]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def _write_funding_sheet(ws, inputs: Inputs) -> None:
    start_year = inputs.starting_year
    forecast_years = inputs.forecast_years
    max_components_rows = int(inputs.features.get("max_components_rows", 500))
    components_end_row = 1 + max_components_rows

    ws.cell(row=1, column=1, value="component_id")
    ws.cell(row=1, column=2, value="component_name")
    for offset in range(forecast_years):
        ws.cell(row=1, column=3 + offset, value=start_year + offset)

    start_year_cell = f"Inputs!$B${INPUT_ROWS['starting_year']}"
    inflation_cell = f"Inputs!$B${INPUT_ROWS['inflation_rate']}"
    inflation_offset = _format_inflation_offset(inputs.spend_inflation_offset)

    for row in range(2, components_end_row + 1):
        ws.cell(row=row, column=1, value=f"=Components!$A{row}")
        ws.cell(row=row, column=2, value=f"=Components!$B{row}")

        base_cost_cell = f"Components!$D{row}"
        spend_year_cell = f"Components!$E{row}"
        recurring_cell = f"Components!$F{row}"
        interval_cell = f"Components!$G{row}"
        include_cell = f"Components!$H{row}"
        id_cell = f"Components!$A{row}"

        for offset in range(forecast_years):
            col = 3 + offset
            year_cell = f"{get_column_letter(col)}$1"

            inflated_cost = (
                f"{base_cost_cell}*(1+{inflation_cell})^"
                f"({year_cell}-{start_year_cell}{inflation_offset})"
            )

            delta = f"({year_cell}-{spend_year_cell})"
            recurring_age = (
                f"IF({interval_cell}<=0,0,"
                f"IF({delta}>=0,"
                f"IF(MOD({delta},{interval_cell})=0,{interval_cell},"
                f"MOD({delta},{interval_cell})),"
                f"{interval_cell}+{delta}))"
            )
            recurring_fraction = (
                f"IF({interval_cell}<=0,0,"
                f"IF({recurring_age}<=0,0,({recurring_age})/{interval_cell}))"
            )

            nonrecurring_fraction = (
                f"IF({year_cell}>{spend_year_cell},0,"
                f"IF({spend_year_cell}-{start_year_cell}<=0,1,"
                f"({year_cell}-{start_year_cell})/"
                f"({spend_year_cell}-{start_year_cell})))"
            )

            funded = (
                f"=IF({id_cell}=\"\",0,"
                f"IF({include_cell}<>\"Y\",0,"
                f"{inflated_cost}*IF({recurring_cell}=\"Y\","
                f"{recurring_fraction},{nonrecurring_fraction})))"
            )
            ws.cell(row=row, column=col, value=funded).number_format = "#,##0"

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 32


def _write_schedule_sheet(ws, inputs: Inputs, schedule_items: List[ScheduleItem]) -> None:
    headers = ["year", "component_id", "component_name", "nominal_expense"]
    ws.append(headers)
    _style_header(ws, len(headers))

    start_row = 2
    inflation_cell = f"Inputs!$B${INPUT_ROWS['inflation_rate']}"
    start_year_cell = f"Inputs!$B${INPUT_ROWS['starting_year']}"
    # spend_inflation_offset models timing (start/mid/end of year) in the exponent.
    inflation_offset = _format_inflation_offset(inputs.spend_inflation_offset)

    for idx, item in enumerate(schedule_items, start=start_row):
        ws.cell(row=idx, column=1, value=item.year)
        ws.cell(row=idx, column=2, value=item.component_id)
        ws.cell(row=idx, column=3, value=item.component_name)

        base_cost_cell = f"Components!$D${item.component_row}"
        formula = (
            f"={base_cost_cell}*(1+{inflation_cell})^"
            f"(A{idx}-{start_year_cell}{inflation_offset})"
        )
        cell = ws.cell(row=idx, column=4, value=formula)
        cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    widths = [8, 16, 32, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def _fully_funded_balance(
    components: Iterable[Component],
    inputs: Inputs,
    year: int,
) -> float:
    total = 0.0
    for component in components:
        if not component.include:
            continue

        inflated_cost = component.base_cost * (
            (1 + inputs.inflation_rate)
            ** (year - inputs.starting_year + inputs.spend_inflation_offset)
        )

        if component.recurring:
            interval = component.interval_years or 0
            if interval <= 0:
                fraction = 0.0
            else:
                if year <= component.spend_year:
                    years_to_next = component.spend_year - year
                else:
                    cycles = math.ceil((year - component.spend_year) / interval)
                    years_to_next = component.spend_year + cycles * interval - year
                recurring_age = interval - years_to_next
                fraction = 0.0 if recurring_age <= 0 else recurring_age / interval
        else:
            if year > component.spend_year:
                fraction = 0.0
            elif component.spend_year - inputs.starting_year <= 0:
                fraction = 1.0
            else:
                fraction = (year - inputs.starting_year) / (
                    component.spend_year - inputs.starting_year
                )

        total += inflated_cost * fraction

    return total


def _compute_audit_expected(
    inputs: Inputs,
    components: List[Component],
    schedule_items: List[ScheduleItem],
    contributions: Dict[int, float],
) -> List[List[float | None]]:
    schedule_expenses = expenses_by_year(schedule_items)
    forecast_rows = compute_forecast(inputs, contributions, schedule_expenses)
    end_year = inputs.starting_year + inputs.forecast_years - 1

    cumulative_contrib = 0.0
    cumulative_interest = 0.0
    expected_rows: List[List[float | None]] = []

    for row in forecast_rows:
        cumulative_contrib += row.contributions
        cumulative_interest += row.interest

        fully_funded = _fully_funded_balance(components, inputs, row.year)
        percent_funded = None if fully_funded == 0 else row.begin_balance / fully_funded

        coverage_end = min(row.year + 4, end_year)
        coverage_sum = sum(
            schedule_expenses.get(year, 0.0)
            for year in range(row.year, coverage_end + 1)
        )
        coverage_5yr = None if coverage_sum == 0 else row.begin_balance / coverage_sum

        expected_rows.append([
            float(row.year),
            row.begin_balance,
            row.contributions,
            cumulative_contrib,
            row.interest,
            cumulative_interest,
            row.expenses,
            row.end_balance,
            percent_funded,
            coverage_5yr,
        ])

    return expected_rows


def _write_forecast_sheet(
    ws,
    inputs: Inputs,
    components: List[Component],
    schedule_items: List[ScheduleItem],
    contributions: Dict[int, float],
) -> None:
    base_headers = list(FORECAST_HEADERS)
    enable_audit = bool(inputs.features.get("enable_audit", False))
    headers = list(base_headers)
    if enable_audit:
        audit_headers: List[str] = []
        for header in base_headers:
            audit_headers.append(f"{header}_expected")
            audit_headers.append(f"{header}_audit")
        headers.extend(audit_headers)
    ws.append(headers)
    _style_header(ws, len(headers))

    start_year = inputs.starting_year
    forecast_years = inputs.forecast_years
    max_schedule_rows = int(inputs.features.get("max_schedule_rows", 10000))
    max_components_rows = int(inputs.features.get("max_components_rows", 500))
    schedule_end_row = 1 + max_schedule_rows
    components_end_row = 1 + max_components_rows

    expected_rows: List[List[float | None]] = []
    audit_columns: List[tuple[int, int, int, str, str]] = []
    if enable_audit:
        expected_rows = _compute_audit_expected(
            inputs, components, schedule_items, contributions
        )
        # Build-time expected values let us flag formula regressions in the workbook.
        expected_start_col = len(base_headers) + 1
        amount_tol_cell = f"Inputs!$B${INPUT_ROWS['audit_tolerance_amount']}"
        ratio_tol_cell = f"Inputs!$B${INPUT_ROWS['audit_tolerance_ratio']}"

        for idx, header in enumerate(base_headers):
            expected_col = expected_start_col + idx * 2
            flag_col = expected_col + 1
            if header == "year":
                tol_cell = "0"
            elif header in AUDIT_RATIO_HEADERS:
                tol_cell = ratio_tol_cell
            else:
                tol_cell = amount_tol_cell
            audit_columns.append((idx + 1, expected_col, flag_col, header, tol_cell))

    for offset in range(forecast_years):
        row = 2 + offset
        year = start_year + offset
        ws.cell(row=row, column=1, value=year)

        if row == 2:
            begin_formula = f"=Inputs!$B${INPUT_ROWS['beginning_reserve_balance']}"
        else:
            begin_formula = f"=H{row - 1}"
        ws.cell(row=row, column=2, value=begin_formula)

        contribution_value = contributions.get(year, 0.0)
        contrib_cell = ws.cell(row=row, column=3, value=contribution_value)
        contrib_cell.fill = INPUT_FILL

        cumulative_contrib_formula = (
            f"=C{row}" if row == 2 else f"=D{row - 1}+C{row}"
        )
        ws.cell(row=row, column=4, value=cumulative_contrib_formula)

        interest_formula = f"=B{row}*Inputs!$B${INPUT_ROWS['investment_return_rate']}"
        ws.cell(row=row, column=5, value=interest_formula)

        cumulative_interest_formula = (
            f"=E{row}" if row == 2 else f"=F{row - 1}+E{row}"
        )
        ws.cell(row=row, column=6, value=cumulative_interest_formula)

        expenses_formula = (
            f"=SUMIFS(Schedule!$D$2:$D${schedule_end_row},"
            f"Schedule!$A$2:$A${schedule_end_row},A{row})"
        )
        ws.cell(row=row, column=7, value=expenses_formula)

        end_formula = f"=B{row}+C{row}+E{row}-G{row}"
        ws.cell(row=row, column=8, value=end_formula)

        funding_col_letter = get_column_letter(3 + offset)
        fully_funded = (
            f"SUM(Funding!${funding_col_letter}$2:${funding_col_letter}${components_end_row})"
        )
        percent_funded_formula = f"=IF({fully_funded}=0,\"\",B{row}/({fully_funded}))"
        ws.cell(row=row, column=9, value=percent_funded_formula)

        coverage_sum = (
            f"SUM(G{row}:INDEX($G$2:$G${1 + forecast_years},"
            f"MIN({row}+4,{1 + forecast_years})-1))"
        )
        coverage_formula = f"=IF({coverage_sum}=0,\"\",B{row}/{coverage_sum})"
        ws.cell(row=row, column=10, value=coverage_formula)

        if enable_audit:
            expected_values = expected_rows[offset]
            for idx, (actual_col, expected_col, flag_col, header, tol_cell) in enumerate(
                audit_columns
            ):
                expected_value = expected_values[idx]
                expected_cell = ws.cell(
                    row=row,
                    column=expected_col,
                    value=expected_value,
                )
                if header == "year":
                    expected_cell.number_format = "0"
                elif header == "percent_funded":
                    expected_cell.number_format = "0.00%"
                elif header == "coverage_5yr":
                    expected_cell.number_format = "0.00"
                else:
                    expected_cell.number_format = "#,##0"

                actual_ref = f"{get_column_letter(actual_col)}{row}"
                expected_ref = f"{get_column_letter(expected_col)}{row}"
                flag_formula = (
                    f"=IF(OR({actual_ref}=\"\",{expected_ref}=\"\"),\"\","
                    f"IF(ABS({actual_ref}-{expected_ref})>{tol_cell},\"FAIL\",\"\"))"
                )
                audit_cell = ws.cell(row=row, column=flag_col, value=flag_formula)
                audit_cell.number_format = "@"

    for row in range(2, 2 + forecast_years):
        for col in range(2, 9):
            ws.cell(row=row, column=col).number_format = "#,##0"
        ws.cell(row=row, column=9).number_format = "0.00%"
        ws.cell(row=row, column=10).number_format = "0.00"

    if enable_audit:
        end_row = 1 + forecast_years
        for actual_col, expected_col, flag_col, _, _ in audit_columns:
            expected_letter = get_column_letter(expected_col)
            ws.column_dimensions[expected_letter].hidden = True
            ws.column_dimensions[get_column_letter(flag_col)].width = 8

    ws.freeze_panes = "A2"
    widths = [8, 18, 16, 22, 16, 20, 16, 18, 16, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def _write_checks_sheet(ws, inputs: Inputs) -> None:
    forecast_years = inputs.forecast_years
    end_row = 1 + forecast_years

    ws.append(["Check", "Value"])
    _style_header(ws, 2)

    ws.append([
        "Years with negative ending balance",
        f"=COUNTIF(Forecast!$H$2:$H${end_row},\"<0\")",
    ])
    ws.append([
        "Years with zero expenses",
        f"=COUNTIF(Forecast!$G$2:$G${end_row},0)",
    ])
    ws.append([
        "Years with coverage_5yr < 1.0",
        f"=COUNTIF(Forecast!$J$2:$J${end_row},\"<1\")",
    ])
    ws.append([
        "Years with coverage_5yr < 0.5",
        f"=COUNTIF(Forecast!$J$2:$J${end_row},\"<0.5\")",
    ])

    ws.append(["", ""])
    ws.append(["Negative balance years", ""])
    ws.append(["Year", ""])

    start_list_row = ws.max_row + 1
    for offset in range(forecast_years):
        forecast_row = 2 + offset
        formula = f"=IF(Forecast!$H${forecast_row}<0,Forecast!$A${forecast_row},\"\")"
        ws.cell(row=start_list_row + offset, column=1, value=formula)

    if inputs.features.get("enable_audit", False):
        ws.append(["", ""])
        ws.append(["Audit summary", ""])
        expected_start_col = len(FORECAST_HEADERS) + 1
        for idx, header in enumerate(FORECAST_HEADERS):
            flag_col = expected_start_col + idx * 2 + 1
            flag_letter = get_column_letter(flag_col)
            ws.append([
                f"Audit flags: {header}",
                f"=COUNTIF(Forecast!${flag_letter}$2:${flag_letter}${end_row},\"FAIL\")",
            ])

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"


def _write_dashboard_sheet(ws, inputs: Inputs) -> None:
    forecast_years = inputs.forecast_years
    end_row = 1 + forecast_years

    ws.append(["Metric", "Value"])
    _style_header(ws, 2)

    ws.append([
        "Lowest reserve balance",
        f"=MIN(Forecast!$H$2:$H${end_row})",
    ])
    ws.append([
        "Ending balance (final year)",
        f"=Forecast!$H${end_row}",
    ])
    ws.append([
        "Final forecast year",
        f"=Forecast!$A${end_row}",
    ])

    for row in range(2, 5):
        ws.cell(row=row, column=2).number_format = "#,##0"
    ws.cell(row=4, column=2).number_format = "0"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"
