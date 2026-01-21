from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List

import yaml
from openpyxl import load_workbook

from .build import build_workbook
from .constants import DATA_DIR
from .model import compute_forecast, expenses_by_year
from .schedule import expand_schedule
from .validate import ValidationError, validate_scenario


@dataclass(frozen=True)
class Fixture:
    name: str
    data_dir: Path
    expected_path: Path
    scenario: str


@dataclass(frozen=True)
class FixtureRunResult:
    issues: List[str]
    output_path: Path | None = None


def load_fixture(expected_path: Path) -> Fixture:
    data_dir = expected_path.parent
    expected = _load_expected(expected_path)
    scenario = expected.get("scenario")
    if not scenario:
        raise ValueError(f"Missing scenario in {expected_path}")
    return Fixture(
        name=data_dir.name,
        data_dir=data_dir,
        expected_path=expected_path,
        scenario=str(scenario),
    )


def find_fixtures(fixtures_root: Path | None = None) -> List[Fixture]:
    root = fixtures_root or (DATA_DIR / "fixtures")
    return [load_fixture(path) for path in root.rglob("expected_values.yaml")]


def run_fixture(fixture: Fixture) -> FixtureRunResult:
    issues: List[str] = []
    expected = _load_expected(fixture.expected_path)

    expected_validation = expected.get("expect", {}).get("validation", {})
    expected_errors = _as_list(expected_validation.get("errors", []))
    expected_warnings = _as_list(expected_validation.get("warnings", []))

    result, inputs, components, contributions = validate_scenario(
        fixture.scenario, data_dir=fixture.data_dir
    )
    issues.extend(
        _compare_lists("validation errors", expected_errors, result.errors)
    )
    issues.extend(
        _compare_lists("validation warnings", expected_warnings, result.warnings)
    )

    if expected_errors:
        return FixtureRunResult(issues)

    if result.errors:
        issues.append("Validation produced errors but none were expected.")
        return FixtureRunResult(issues)

    try:
        output_path, build_result = build_workbook(
            fixture.scenario, data_dir=fixture.data_dir
        )
    except ValidationError as exc:
        issues.append(f"Build failed: {exc}")
        issues.extend(exc.result.errors)
        return FixtureRunResult(issues)

    issues.extend(
        _compare_lists(
            "build validation warnings", expected_warnings, build_result.warnings
        )
    )

    workbook_expect = expected.get("expect", {}).get("workbook", {})
    model_expect = expected.get("expect", {}).get("model", {})

    if workbook_expect:
        issues.extend(_check_workbook(output_path, workbook_expect))
    if model_expect:
        issues.extend(
            _check_model(inputs, components, contributions, model_expect)
        )

    return FixtureRunResult(issues, output_path)


def _load_expected(path: Path) -> Dict[str, Any]:
    data = yaml.safe_load(path.read_text()) or {}
    if not isinstance(data, dict):
        raise ValueError(f"Invalid expected_values.yaml in {path}")
    return data


def _as_list(value: Iterable[Any]) -> List[str]:
    return [str(item) for item in value]


def _compare_lists(label: str, expected: List[str], actual: List[str]) -> List[str]:
    if sorted(expected) == sorted(actual):
        return []
    return [
        f"Mismatch in {label}. expected={sorted(expected)} actual={sorted(actual)}"
    ]


def _check_workbook(path: Path, expect: Dict[str, Any]) -> List[str]:
    issues: List[str] = []
    wb = load_workbook(path, data_only=False)

    values = expect.get("values", [])
    formulas = expect.get("formulas", [])

    for entry in values:
        issues.extend(_check_cell_value(wb, entry))

    for entry in formulas:
        issues.extend(_check_cell_formula(wb, entry))

    return issues


def _check_cell_value(wb, entry: Dict[str, Any]) -> List[str]:
    issues: List[str] = []
    sheet = entry.get("sheet")
    cell = entry.get("cell")
    expected = entry.get("equals")
    tolerance = entry.get("tolerance", 1e-6)

    if sheet not in wb.sheetnames:
        return [f"Missing sheet: {sheet}"]

    ws = wb[sheet]
    actual = ws[cell].value

    if _is_number(expected) and _is_number(actual):
        if abs(float(actual) - float(expected)) > float(tolerance):
            issues.append(
                f"{sheet}!{cell} expected {expected} got {actual}"
            )
    else:
        if actual != expected:
            issues.append(
                f"{sheet}!{cell} expected {expected} got {actual}"
            )

    return issues


def _check_cell_formula(wb, entry: Dict[str, Any]) -> List[str]:
    issues: List[str] = []
    sheet = entry.get("sheet")
    cell = entry.get("cell")
    expected = entry.get("equals")

    if sheet not in wb.sheetnames:
        return [f"Missing sheet: {sheet}"]

    ws = wb[sheet]
    actual = ws[cell].value
    if actual != expected:
        issues.append(
            f"{sheet}!{cell} formula expected {expected} got {actual}"
        )

    return issues


def _check_model(
    inputs, components, contributions, expect: Dict[str, Any]
) -> List[str]:
    issues: List[str] = []

    if inputs.features.get("enable_schedule_expansion", True):
        schedule_items = expand_schedule(components, inputs)
    else:
        schedule_items = []

    expenses = expenses_by_year(schedule_items)
    forecast_rows = compute_forecast(inputs, contributions, expenses)

    schedule_expect = expect.get("schedule", [])
    forecast_expect = expect.get("forecast", [])
    counts_expect = expect.get("counts", {})

    schedule_map = {
        (item.year, item.component_id): item.nominal_expense
        for item in schedule_items
    }
    forecast_map = {row.year: row for row in forecast_rows}

    for entry in schedule_expect:
        year = int(entry["year"])
        component_id = entry["component_id"]
        expected = float(entry["nominal_expense"])
        tolerance = float(entry.get("tolerance", 1e-6))
        actual = schedule_map.get((year, component_id))
        if actual is None:
            issues.append(
                f"Schedule missing {component_id} in {year}"
            )
            continue
        if abs(actual - expected) > tolerance:
            issues.append(
                f"Schedule {component_id} {year} expected {expected} got {actual}"
            )

    for entry in forecast_expect:
        year = int(entry["year"])
        row = forecast_map.get(year)
        if row is None:
            issues.append(f"Forecast missing year {year}")
            continue
        tolerance = float(entry.get("tolerance", 1e-6))
        for field in ("begin_balance", "contributions", "interest", "expenses", "end_balance"):
            if field not in entry:
                continue
            expected = float(entry[field])
            actual = float(getattr(row, field))
            if abs(actual - expected) > tolerance:
                issues.append(
                    f"Forecast {year} {field} expected {expected} got {actual}"
                )

    if counts_expect:
        schedule_count = len(schedule_items)
        negative_count = sum(1 for row in forecast_rows if row.end_balance < 0)
        zero_expense_count = sum(1 for row in forecast_rows if row.expenses == 0)
        forecast_years = len(forecast_rows)

        counts_actual = {
            "schedule_items": schedule_count,
            "negative_balance_years": negative_count,
            "zero_expense_years": zero_expense_count,
            "forecast_years": forecast_years,
        }

        for key, expected in counts_expect.items():
            actual = counts_actual.get(key)
            if actual != expected:
                issues.append(
                    f"Count {key} expected {expected} got {actual}"
                )

    return issues


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float))
