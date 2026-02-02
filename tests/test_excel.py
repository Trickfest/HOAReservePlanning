import unittest

from openpyxl.utils import get_column_letter

from reserve import excel
from reserve.model import Component, Inputs
from reserve.schedule import expand_schedule


def _expected_percent_funded_formula(
    row: int,
    max_components_rows: int,
) -> str:
    components_end_row = 1 + max_components_rows
    # Forecast rows start at 2, Funding years start at column C (3).
    funding_col = get_column_letter(3 + (row - 2))
    fully_funded = (
        f"SUM(Funding!${funding_col}$2:${funding_col}${components_end_row})"
    )
    return f"=IF({fully_funded}=0,\"\",B{row}/({fully_funded}))"


def _expected_coverage_formula(row: int, forecast_years: int) -> str:
    end_row = 1 + forecast_years
    coverage_sum = (
        f"SUM(G{row}:INDEX($G$2:$G${end_row},MIN({row}+4,{end_row})-1))"
    )
    return f"=IF({coverage_sum}=0,\"\",B{row}/{coverage_sum})"


def _expected_expenses_formula(row: int, max_schedule_rows: int) -> str:
    schedule_end_row = 1 + max_schedule_rows
    return (
        f"=SUMIFS(Schedule!$D$2:$D${schedule_end_row},"
        f"Schedule!$A$2:$A${schedule_end_row},A{row})"
    )


class ExcelTests(unittest.TestCase):
    def test_forecast_formulas(self) -> None:
        inputs = Inputs(
            starting_year=2025,
            forecast_years=3,
            beginning_reserve_balance=1000.0,
            inflation_rate=0.02,
            investment_return_rate=0.01,
            features={"max_components_rows": 20, "max_schedule_rows": 50},
        )
        components = [
            Component(
                id="roof",
                name="Roof",
                category="Building",
                base_cost=1000.0,
                spend_year=2026,
                recurring=False,
                interval_years=None,
                include=True,
                row_index=2,
            )
        ]
        schedule_items = expand_schedule(components, inputs)
        contributions = {2025: 0.0, 2026: 0.0, 2027: 0.0}

        wb = excel.build_workbook(
            inputs=inputs,
            components=components,
            schedule_items=schedule_items,
            contributions=contributions,
            scenario="test",
        )

        forecast_ws = wb["Forecast"]
        headers = [cell.value for cell in forecast_ws[1]]
        self.assertEqual(
            headers,
            [
                "year",
                "begin_balance",
                "contributions",
                "cumulative_contributions",
                "interest",
                "cumulative_interest",
                "expenses",
                "end_balance",
                "percent_funded",
                "coverage_5yr",
            ],
        )

        self.assertEqual(
            forecast_ws["D2"].value,
            "=C2",
        )
        self.assertEqual(
            forecast_ws["D3"].value,
            "=D2+C3",
        )
        self.assertEqual(
            forecast_ws["F2"].value,
            "=E2",
        )
        self.assertEqual(
            forecast_ws["F3"].value,
            "=F2+E3",
        )
        self.assertEqual(
            forecast_ws["G2"].value,
            _expected_expenses_formula(2, inputs.features["max_schedule_rows"]),
        )
        self.assertEqual(
            forecast_ws["I2"].value,
            _expected_percent_funded_formula(
                2,
                inputs.features["max_components_rows"],
            ),
        )
        self.assertEqual(
            forecast_ws["J2"].value,
            _expected_coverage_formula(2, inputs.forecast_years),
        )

        self.assertEqual(forecast_ws["I2"].number_format, "0.00%")
        self.assertEqual(forecast_ws["J2"].number_format, "0.00")

    def test_schedule_formula_spend_inflation_timing_variants(self) -> None:
        variants = [
            ("start_of_year", 0.0, ""),
            ("mid_year", 0.5, "+0.5"),
            ("end_of_year", 1.0, "+1"),
        ]

        for timing, offset, suffix in variants:
            inputs = Inputs(
                starting_year=2025,
                forecast_years=2,
                beginning_reserve_balance=0.0,
                inflation_rate=0.02,
                investment_return_rate=0.0,
                features={"max_components_rows": 10, "max_schedule_rows": 20},
                spend_inflation_timing=timing,
                spend_inflation_offset=offset,
            )
            components = [
                Component(
                    id="roof",
                    name="Roof",
                    category="Building",
                    base_cost=1000.0,
                    spend_year=2026,
                    recurring=False,
                    interval_years=None,
                    include=True,
                    row_index=2,
                )
            ]
            schedule_items = expand_schedule(components, inputs)
            contributions = {2025: 0.0, 2026: 0.0}

            wb = excel.build_workbook(
                inputs=inputs,
                components=components,
                schedule_items=schedule_items,
                contributions=contributions,
                scenario="timing-test",
            )

            schedule_ws = wb["Schedule"]
            expected = (
                "=Components!$D$2*(1+Inputs!$B$4)^(A2-Inputs!$B$2" + suffix + ")"
            )
            self.assertEqual(schedule_ws["D2"].value, expected)

    def test_forecast_audit_enabled(self) -> None:
        inputs = Inputs(
            starting_year=2025,
            forecast_years=2,
            beginning_reserve_balance=1000.0,
            inflation_rate=0.0,
            investment_return_rate=0.0,
            features={
                "enable_audit": True,
                "max_components_rows": 5,
                "max_schedule_rows": 20,
            },
        )
        components = [
            Component(
                id="roof",
                name="Roof",
                category="Building",
                base_cost=1000.0,
                spend_year=2026,
                recurring=False,
                interval_years=None,
                include=True,
                row_index=2,
            )
        ]
        schedule_items = expand_schedule(components, inputs)
        contributions = {2025: 0.0, 2026: 0.0}

        wb = excel.build_workbook(
            inputs=inputs,
            components=components,
            schedule_items=schedule_items,
            contributions=contributions,
            scenario="audit-on",
        )

        forecast_ws = wb["Forecast"]
        headers = [cell.value for cell in forecast_ws[1]]
        base_len = len(excel.FORECAST_HEADERS)
        self.assertEqual(headers[:base_len], excel.FORECAST_HEADERS)
        self.assertEqual(len(headers), base_len * 3)
        self.assertIn("year_expected", headers)
        self.assertIn("year_audit", headers)

        first_expected_col = get_column_letter(base_len + 1)
        first_audit_col = get_column_letter(base_len + 2)
        self.assertTrue(forecast_ws.column_dimensions[first_expected_col].hidden)
        self.assertFalse(forecast_ws.column_dimensions[first_audit_col].hidden)
        self.assertTrue(str(forecast_ws[f"{first_audit_col}2"].value).startswith("=IF("))

        checks_ws = wb["Checks"]
        labels = [cell.value for cell in checks_ws["A"] if cell.value]
        self.assertIn("Audit summary", labels)
        self.assertIn("Audit flags: year", labels)

    def test_forecast_audit_disabled(self) -> None:
        inputs = Inputs(
            starting_year=2025,
            forecast_years=2,
            beginning_reserve_balance=1000.0,
            inflation_rate=0.0,
            investment_return_rate=0.0,
            features={"max_components_rows": 5, "max_schedule_rows": 20},
        )
        components = []
        schedule_items = expand_schedule(components, inputs)
        contributions = {2025: 0.0, 2026: 0.0}

        wb = excel.build_workbook(
            inputs=inputs,
            components=components,
            schedule_items=schedule_items,
            contributions=contributions,
            scenario="audit-off",
        )

        forecast_ws = wb["Forecast"]
        headers = [cell.value for cell in forecast_ws[1]]
        self.assertEqual(headers, excel.FORECAST_HEADERS)
        self.assertEqual(len(forecast_ws.conditional_formatting), 0)

        checks_ws = wb["Checks"]
        labels = [cell.value for cell in checks_ws["A"] if cell.value]
        self.assertNotIn("Audit summary", labels)


if __name__ == "__main__":
    unittest.main()
